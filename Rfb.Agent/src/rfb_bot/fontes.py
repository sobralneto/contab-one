"""De onde vem a lista de clientes a processar.

Duas fontes: os blocos ``[[clientes]]`` do próprio ``config.toml`` (prático
para poucos casos) e uma planilha Excel -- coluna A = CNPJ, coluna B =
Nome --, para quando a lista já existe em uma ferramenta de gestão de
carteira. Trocar de fonte é uma linha em ``config.toml``
(``fonte_clientes``); nenhuma outra parte do robô muda, porque todas
conhecem apenas ``list[Cliente]``.
"""

from __future__ import annotations

import logging
from typing import Callable

from .erros import ErroRobo
from .log import AVISO, obter_logger
from .settings import FONTE_EXCEL, FONTE_TOML, Cliente, Config, normalizar_documento


def _texto_celula(valor: object) -> str:
    """Converte o valor de uma célula para texto, sem sufixo ``.0``.

    Se a coluna do CNPJ não estiver formatada como texto no Excel, o
    openpyxl devolve ``float`` (ex.: ``12345678000199.0``) -- e ``str()``
    direto colaria o ``.0`` no meio dos dígitos ao normalizar.
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip()


def _do_toml(cfg: Config) -> list[Cliente]:
    return list(cfg.clientes)


def _do_excel(cfg: Config) -> list[Cliente]:
    """Lê a lista de clientes de uma planilha: coluna A = CNPJ, coluna B = Nome.

    A primeira linha pode ser cabeçalho (``CNPJ``, ``Nome``, ...) ou já o
    primeiro dado -- é reconhecida como cabeçalho e pulada quando a coluna A
    não contém nenhum dígito. Linhas com a coluna A vazia são ignoradas, para
    tolerar espaçamento na planilha.
    """
    caminho = cfg.caminho_clientes_arquivo()
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencia opcional
        raise ErroRobo(
            "fonte_clientes='excel' exige o pacote 'openpyxl' (pip install openpyxl)."
        ) from exc

    try:
        planilha = load_workbook(caminho, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - arquivo corrompido, senha, etc.
        raise ErroRobo(f"Nao foi possivel abrir a planilha '{caminho}': {exc}") from exc

    clientes: list[Cliente] = []
    try:
        aba = planilha.active
        for numero, linha in enumerate(aba.iter_rows(values_only=True), start=1):
            cnpj_bruto = _texto_celula(linha[0]) if linha else ""
            if not cnpj_bruto:
                continue

            digitos = normalizar_documento(cnpj_bruto)
            if numero == 1 and not digitos:
                continue  # cabeçalho ("CNPJ", "Nome"): sem digito nenhum na coluna A

            if not digitos:
                raise ErroRobo(
                    f"Linha {numero} de '{caminho.name}': coluna A (CNPJ) sem "
                    f"digito utilizavel: {cnpj_bruto!r}"
                )
            # CNPJ digitado como numero no Excel perde zeros a esquerda
            # (comum quando a coluna nao esta formatada como texto).
            # Recompor com zfill e mais seguro do que travar a importacao
            # inteira por causa de um formato de celula.
            if len(digitos) < 14:
                digitos = digitos.zfill(14)

            nome = _texto_celula(linha[1]) if len(linha) > 1 else ""
            # Sem nome na planilha, o CNPJ formatado serve de rotulo -- ruim
            # para o campo "Nome" da credencial, mas nao trava a importacao
            # por uma coluna B vazia.
            clientes.append(Cliente(cnpj=digitos, nome_credencial=nome or digitos))
    finally:
        planilha.close()

    if not clientes:
        raise ErroRobo(f"Nenhum cliente encontrado na planilha '{caminho}'.")
    return clientes


# Registro de leitores por nome de fonte.
_LEITORES: dict[str, Callable[[Config], list[Cliente]]] = {
    FONTE_TOML: _do_toml,
    FONTE_EXCEL: _do_excel,
}


def carregar_clientes(cfg: Config, log: logging.Logger | None = None) -> list[Cliente]:
    """Devolve os clientes da fonte configurada em ``cfg.fonte_clientes``."""
    log = log or obter_logger("fontes")

    if cfg.fonte_clientes == FONTE_EXCEL and cfg.clientes:
        log.warning(
            "%s fonte_clientes='excel': os %d bloco(s) [[clientes]] do "
            "config.toml serao ignorados; a lista vem de '%s'.",
            AVISO, len(cfg.clientes), cfg.clientes_arquivo,
        )

    leitor = _LEITORES.get(cfg.fonte_clientes)
    if leitor is None:
        raise ErroRobo(f"Fonte de clientes '{cfg.fonte_clientes}' sem leitor registrado.")
    clientes = leitor(cfg)
    log.info("Fonte '%s': %d cliente(s) carregado(s).", cfg.fonte_clientes, len(clientes))
    return clientes
