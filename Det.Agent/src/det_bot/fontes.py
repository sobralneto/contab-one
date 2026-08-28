"""De onde vem a lista de empresas a consultar.

Hoje a lista vem de uma planilha Excel fixa -- ``empresas/empresas.xlsx``,
coluna A = CNPJ, coluna B = Nome. Amanhã virá de uma API. Este módulo existe
para que essa troca seja **uma função nova aqui**, e não uma cirurgia no
runner: todo o resto do robô só conhece :func:`carregar_empresas`, que
devolve ``list[Empresa]`` independentemente da origem.

Para plugar a API quando ela existir:

1. escreva ``_da_api(cfg)`` devolvendo ``list[Empresa]``;
2. registre-a em :data:`_LEITORES` sob a chave ``"api"``;
3. acrescente ``"api"`` a ``FONTES_EMPRESAS`` em :mod:`det_bot.settings`;
4. troque ``"fonte_empresas": "api"`` no ``config/empresas.json``.

Nenhuma outra parte do código muda.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Callable

from .erros import ErroRobo
from .log import AVISO, obter_logger
from .settings import (
    FONTE_EXCEL,
    Config,
    Empresa,
    formatar_cnpj,
    normalizar_cnpj,
)


def _texto_celula(valor: object) -> str:
    """Converte o valor de uma célula para texto, sem sufixo ``.0``.

    Se a coluna do CNPJ não estiver formatada como texto no Excel, o
    openpyxl devolve ``float`` (ex.: ``12345678000199.0``) -- e ``str()``
    direto colaria o ``.0`` no meio dos dígitos ao normalizar.
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip()


def _do_excel(cfg: Config) -> list[Empresa]:
    """Lê a lista de empresas de uma planilha: coluna A = CNPJ, coluna B = Nome.

    A primeira linha pode ser cabeçalho (``CNPJ``, ``Nome``, ...) ou já o
    primeiro dado -- é reconhecida como cabeçalho e pulada quando a coluna A
    não contém nenhum dígito. Linhas com a coluna A vazia são ignoradas, para
    tolerar espaçamento na planilha.
    """
    caminho = cfg.caminho_empresas_arquivo()
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencia opcional
        raise ErroRobo(
            "fonte_empresas='excel' exige o pacote 'openpyxl' (pip install "
            "openpyxl)."
        ) from exc

    try:
        planilha = load_workbook(caminho, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - arquivo corrompido, senha, etc.
        raise ErroRobo(f"Nao foi possivel abrir a planilha '{caminho}': {exc}") from exc

    empresas: list[Empresa] = []
    try:
        aba = planilha.active
        for numero, linha in enumerate(aba.iter_rows(values_only=True), start=1):
            cnpj_bruto = _texto_celula(linha[0]) if linha else ""
            if not cnpj_bruto:
                continue

            digitos = normalizar_cnpj(cnpj_bruto)
            if numero == 1 and not digitos:
                continue  # cabeçalho ("CNPJ", "Nome"): sem digito nenhum na coluna A

            if not digitos:
                raise ErroRobo(
                    f"Linha {numero} de '{caminho.name}': coluna A (CNPJ) sem "
                    f"digito utilizavel: {cnpj_bruto!r}"
                )
            # CNPJ digitado como numero no Excel perde zeros a esquerda
            # (comum quando a coluna nao esta formatada como texto).
            # Recompor com zfill e mais seguro do que travar a importacao
            # inteira por causa de um formato de celula.
            if len(digitos) < 14:
                digitos = digitos.zfill(14)

            nome = _texto_celula(linha[1]) if len(linha) > 1 else ""
            empresas.append(Empresa(cnpj=digitos, nome=nome or None, id=digitos))
    finally:
        planilha.close()

    if not empresas:
        raise ErroRobo(f"Nenhuma empresa encontrada na planilha '{caminho}'.")
    return empresas


# Registro de leitores por nome de fonte. Ver a docstring do módulo para
# acrescentar a API.
_LEITORES: dict[str, Callable[[Config], list[Empresa]]] = {
    FONTE_EXCEL: _do_excel,
}


def carregar_empresas(cfg: Config, log: logging.Logger | None = None) -> list[Empresa]:
    """Devolve as empresas da fonte configurada em ``cfg.fonte_empresas``."""
    log = log or obter_logger("fontes")
    leitor = _LEITORES.get(cfg.fonte_empresas)
    if leitor is None:
        raise ErroRobo(
            f"Fonte de empresas '{cfg.fonte_empresas}' sem leitor registrado."
        )
    empresas = leitor(cfg)
    log.info("Fonte '%s': %d empresa(s) carregada(s).", cfg.fonte_empresas, len(empresas))
    return empresas


def carregar_env(raiz: Path) -> None:
    """Carrega o ``.env`` da raiz do projeto, se o python-dotenv existir.

    Valores já presentes no ambiente têm precedência -- assim o agendador ou
    o CI podem sobrepor o arquivo sem editá-lo.
    """
    try:
        from dotenv import load_dotenv
    except ImportError:  # pragma: no cover - dependencia opcional
        return
    load_dotenv(raiz / ".env", override=False)


def montar_configuracao(
    caminho_config: str | Path,
    log: logging.Logger | None = None,
    *,
    empresas_arquivo: str | None = None,
) -> Config:
    """Carrega o JSON, injeta as empresas da fonte e valida o conjunto.

    É o único ponto que costura configuração + segredos + lista de trabalho,
    de modo que o runner receba um ``Config`` já coerente.

    ``empresas_arquivo``, quando informado (via ``--empresas-arquivo`` na
    CLI), força a fonte "excel" com essa planilha -- sobrepondo
    ``fonte_empresas`` do JSON antes da lista ser carregada, já que o
    arquivo recebido muda a cada execução.
    """
    cfg = Config.carregar(caminho_config)
    carregar_env(cfg.raiz)

    if empresas_arquivo:
        cfg.fonte_empresas = FONTE_EXCEL
        cfg.empresas_arquivo = empresas_arquivo

    cfg.empresas = carregar_empresas(cfg, log)
    cfg.aplicar_perfil_padrao()
    cfg.validar()

    if log:
        suspeitos = cfg.cnpjs_suspeitos()
        if suspeitos:
            log.warning(
                "%s CNPJ(s) com digito verificador invalido: %s. O robo vai "
                "consultar assim mesmo -- confira a planilha '%s'.",
                AVISO,
                [formatar_cnpj(e.cnpj) for e in suspeitos],
                cfg.empresas_arquivo,
            )
    return cfg
